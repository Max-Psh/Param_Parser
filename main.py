import openpyxl

def save_params(filename, text, key):
    f = open(filename, key)
    f.write(text)
    f.close()


def header(sheet,shift):
    rows = sheet.max_row
    text=''
    for i in range(rows):
        param_name = sheet.cell(row=i+1, column=1)
        param_val = sheet.cell(row=i+1, column=2)
        text = text + shift  + str(param_name.value) + '=' + str(param_val.value)+'\n'
    return text


def body(sheet,shift):
    num_param = sheet.cell(row=1, column=2).value
    rows = sheet.max_row
    text=''
    for i in range(2,rows):
        text = text + shift + '$'+str(i-2)+'\n'
        for k in range(num_param):
            param_name = sheet.cell(2, column=k+2).value
            param_val = sheet.cell(row=i + 1, column=k+2).value
            text = text + shift + '\t' + str(param_name) + '=' + str(param_val) + '\n'
        text = text + shift + '$'+str(i-2)+'\n'

    return text


def wrapped_body(wb,wrapper,HDR_sheet, Data_sheets, shift):
    shift = shift
    text = '\n'+ shift + wrapper + '\n'
    if HDR_sheet:
        text = text + '\n' + header(wb[HDR_sheet], shift + '\t')
    for sheet in Data_sheets:
        if sheet:
            text = text + '\n' + body(wb[sheet], shift + '\t')
    text = text + '\n' + shift + wrapper +'\n'
    return text


def system_block(wb):
    shift = '\t'
    text ='$SYSTEM\n'
    text = text + '\n' + header(wb['Main System Params'], shift)
    text = text + wrapped_body(wb, '$PCI', 'System PCI header', ['System PCI Regs'], shift)
    text = text+'\n$SYSTEM\n'
    return text


def tables_block(wb):
    text ='\n$TABLES\n'
    text = text+'\n$TABLES\n'
    return text


def params_block(wb):
    text ='\n$PARAMS\n'
    text = text + '\n' + body(wb['Params'], '\t')
    text = text+'\n$PARAMS\n'
    return text


def LOG_block(wb):
    shift = '\t'
    text ='\n$LOG\n'
    text = text + '\n' + header(wb['LOG Header'], shift)
    text = text + '\n' + body(wb['LOG Param'], shift)
    text = text+'\n$LOG\n'
    return text


def OSC_block(wb):
    shift = '\t'
    text ='\n$OSCILLOSCOPE\n'
    text = text + '\n' + header(wb['OSC Header'], shift)
    text =text + wrapped_body(wb, '$DATA', '',['OSC Data'], shift)
    text = text+'\n$OSCILLOSCOPE\n'
    return text


def Modbus_block(wb):
    shift = '\t'
    text ='\n$MODBUS\n'
    text = text + '\n' + header(wb['MB Header'], shift)
    modbus_sheets = {
                    'MB Coils':'COILS',
        'MB Input discretes':'INPUT_DISCRETES',
        'MB holdings registers':'HOLDINGS_REGISTERS',
        'MB input registers': 'INPUT_REGISTERS'
                   }
    for sheet in modbus_sheets:
        text = text + wrapped_body(wb, '$'+modbus_sheets[sheet], '',[sheet], shift)
    text = text+'\n$MODBUS\n'
    return text

def IEC60870_block(wb):
    shift = '\t'
    text ='\n$IEC60870\n'
    text = text + '\n' + header(wb['104 Header'], shift)
    text =text + wrapped_body(wb, '$IN', '',['104 In Regs'], shift)
    text = text + wrapped_body(wb, '$OUT', '', ['104 Out Regs'], shift)
    text = text+'\n$IEC60870\n'
    return text

def Strings_block(wb):
    sheet = wb['Strings']
    shift = '\t'
    text ='\n$STRINGS\n'
    rows = sheet.max_row
    for i in range(1,rows):
        param_num = sheet.cell(row=i + 1, column=2).value
        param_val = sheet.cell(row=i + 1, column=3).value
        text = text + shift + '\t' + str(param_num) + '=' + str(param_val) + '\n'
    text = text+'\n$STRINGS\n'
    return text

def main():

    wb = openpyxl.load_workbook('Params.xlsx')

    filename ='Param.conf'
    save_params(filename, system_block(wb), 'w')
    save_params(filename, tables_block(wb), 'a')
    save_params(filename, params_block(wb), 'a')
    save_params(filename, LOG_block(wb), 'a')
    save_params(filename, OSC_block(wb), 'a')
    save_params(filename, Modbus_block(wb), 'a')
    save_params(filename, IEC60870_block(wb), 'a')
    save_params(filename, Strings_block(wb), 'a')



if __name__ == '__main__':
    main()

